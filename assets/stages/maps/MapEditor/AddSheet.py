import os
from openpyxl import Workbook, load_workbook

def add_sheet_to_excel(file_name, sheet_name):
    # 检查文件是否存在
    if not os.path.exists(file_name):
        print(f"文件 {file_name} 不存在")
        return

    # 尝试打开Excel文件
    try:
        workbook = load_workbook(file_name)
    except Exception as e:
        print(f"打开文件时出错: {e}")
        return

    # 检查工作表是否已存在
    if sheet_name in workbook.sheetnames:
        print(f"工作表 '{sheet_name}' 已存在")
    else:
        # 创建新的工作表
        workbook.create_sheet(title=sheet_name)
        print(f"工作表 '{sheet_name}' 已添加")

    # 保存文件
    try:
        workbook.save(file_name)
        print(f"文件 '{file_name}' 已更新")
    except Exception as e:
        print(f"保存文件时出错: {e}")

# 使用示例
file_name = input() + ".xlsx"  # 你的Excel文件名
sheet_name = 'backgroundObjects'    # 你想要添加的工作表名
this_path = os.path.dirname(__file__)

add_sheet_to_excel(os.path.join(this_path, file_name), sheet_name)
